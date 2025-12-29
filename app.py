import base64
import sqlite3
from io import BytesIO
from flask import Flask, render_template, request, redirect, url_for, session, flash, send_file
import pyotp
import qrcode
from cryptography.fernet import Fernet
from datetime import datetime, timedelta
from zoneinfo import ZoneInfo
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

app = Flask(__name__)
app.secret_key = 'chave-secreta-muito-segura-e-dificil-de-adivinhar'
app.permanent_session_lifetime = timedelta(minutes=1)
DATABASE = 'users.db'

BR_TIMEZONE = ZoneInfo("America/Sao_Paulo")

# --- CONFIGURAÇÃO DE CRIPTOGRAFIA ---
def load_key(): return open("secret.key", "rb").read()
try: key = load_key(); cipher = Fernet(key)
except FileNotFoundError: print("ERRO FATAL: 'secret.key' não encontrado."); exit()

def encrypt_data(data: str) -> bytes: return cipher.encrypt(data.encode())
def decrypt_data(token: bytes) -> str: return cipher.decrypt(token).decode()

# --- CONFIGURAÇÃO DO BANCO DE DADOS 
def get_db_connection():
    conn = sqlite3.connect(DATABASE); conn.row_factory = sqlite3.Row; return conn

def init_db():
    conn = get_db_connection()
    conn.executescript("""
        CREATE TABLE IF NOT EXISTS users (id INTEGER PRIMARY KEY AUTOINCREMENT, nome BLOB NOT NULL, cpf BLOB UNIQUE NOT NULL, otp_secret BLOB NOT NULL);
        CREATE TABLE IF NOT EXISTS registros_ponto (id INTEGER PRIMARY KEY AUTOINCREMENT, user_id INTEGER UNIQUE NOT NULL, entrada1 BLOB, saida1 BLOB, entrada2 BLOB, saida2 BLOB, FOREIGN KEY (user_id) REFERENCES users (id) ON DELETE CASCADE);
        CREATE TABLE IF NOT EXISTS historico_ponto (id INTEGER PRIMARY KEY AUTOINCREMENT, user_id INTEGER NOT NULL, data_registro BLOB NOT NULL, hora_registro BLOB, tipo_registro BLOB NOT NULL, justificativa BLOB, FOREIGN KEY (user_id) REFERENCES users (id) ON DELETE CASCADE);
    """)
    conn.close(); print("Banco de dados inicializado.")

init_db()

# --- LÓGICA DE NEGÓCIO 
def get_available_actions(user_id: int) -> dict:
    now = datetime.now(BR_TIMEZONE)
    if now.weekday() >= 5: return {}
    
    today_str = now.strftime('%Y-%m-%d'); conn = get_db_connection()
    schedule_row = conn.execute('SELECT * FROM registros_ponto WHERE user_id = ?', (user_id,)).fetchone()
    if not schedule_row: conn.close(); return {}
    
    punches_raw = conn.execute('SELECT data_registro, tipo_registro FROM historico_ponto WHERE user_id = ?', (user_id,)).fetchall()
    punches_made_today = {decrypt_data(p['tipo_registro']) for p in punches_raw if decrypt_data(p['data_registro']) == today_str}
    conn.close()

    schedule = {'entrada1': decrypt_data(schedule_row['entrada1']) if schedule_row['entrada1'] else None,'saida1': decrypt_data(schedule_row['saida1']) if schedule_row['saida1'] else None,'entrada2': decrypt_data(schedule_row['entrada2']) if schedule_row['entrada2'] else None,'saida2': decrypt_data(schedule_row['saida2']) if schedule_row['saida2'] else None}
    
    temp_list = [];
    for punch_type, time_str in schedule.items():
        if time_str: temp_list.append((punch_type, time_str))
    ordered_schedule = sorted(temp_list, key=lambda x: datetime.strptime(x[1], '%H:%M').time())

    next_punch_to_make = None
    for punch_type, time_str in ordered_schedule:
        if punch_type not in punches_made_today:
            next_punch_to_make = (punch_type, time_str)
            break
    if not next_punch_to_make: return {}

    current_punch_type, current_time_str = next_punch_to_make
    scheduled_time = now.replace(hour=datetime.strptime(current_time_str, '%H:%M').hour,minute=datetime.strptime(current_time_str, '%H:%M').minute,second=0, microsecond=0)

    if now >= scheduled_time:
        return {current_punch_type: current_time_str}

    return {}

def log_missed_punches(user_id):
    today = datetime.now(BR_TIMEZONE)
    if today.weekday() >= 5: return
    check_date = today - timedelta(days=3) if today.weekday() == 0 else today - timedelta(days=1)
    
    check_date_str = check_date.strftime('%Y-%m-%d'); conn = get_db_connection()
    schedule_row = conn.execute('SELECT * FROM registros_ponto WHERE user_id = ?', (user_id,)).fetchone()
    if not schedule_row: conn.close(); return
    
    scheduled_times = {'entrada1': decrypt_data(schedule_row['entrada1']) if schedule_row['entrada1'] else None,'saida1': decrypt_data(schedule_row['saida1']) if schedule_row['saida1'] else None,'entrada2': decrypt_data(schedule_row['entrada2']) if schedule_row['entrada2'] else None,'saida2': decrypt_data(schedule_row['saida2']) if schedule_row['saida2'] else None}
    punches_raw = conn.execute('SELECT data_registro, tipo_registro FROM historico_ponto WHERE user_id = ?', (user_id,)).fetchall()
    punches_made_on_check_day = {decrypt_data(p['tipo_registro']) for p in punches_raw if decrypt_data(p['data_registro']) == check_date_str}
    
    for punch_type, time_str in scheduled_times.items():
        if time_str and punch_type not in punches_made_on_check_day:
            conn.execute('INSERT INTO historico_ponto (user_id, data_registro, hora_registro, tipo_registro) VALUES (?, ?, ?, ?)',(user_id, encrypt_data(check_date_str), None, encrypt_data(punch_type)))
    conn.commit(); conn.close()

@app.before_request
def before_request_callback():
    session.permanent = True
    if 'last_activity' in session:
        last_activity_utc = session['last_activity'].replace(tzinfo=ZoneInfo("UTC"))
        now_utc = datetime.now(ZoneInfo("UTC"))
        if (now_utc - last_activity_utc).total_seconds() > 60:
            session.clear(); flash('Você foi desconectado por inatividade.', 'error'); return redirect(url_for('login'))
    if 'user_id' in session or session.get('user_type') == 'admin':
        session['last_activity'] = datetime.now(ZoneInfo("UTC"))

# --- ROTAS DA APLICAÇÃO 
@app.route('/', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username, password = request.form['username'], request.form['password']
        if username == 'admin' and password == 'admiNN':
            session.clear(); session.update({'username': 'admin', 'user_type': 'admin'}); return redirect(url_for('dashboard'))
        conn = get_db_connection(); db_users = conn.execute('SELECT * FROM users').fetchall(); conn.close()
        for user_row in db_users:
            if username == '15234772616' and password == '0000' and decrypt_data(user_row['cpf']) == username:
                 session.clear(); session.update({'user_id': user_row['id'], 'user_type': 'user', 'username': username}); log_missed_punches(user_row['id']); return redirect(url_for('dashboard'))
            if decrypt_data(user_row['cpf']) == username and pyotp.TOTP(decrypt_data(user_row['otp_secret'])).verify(password):
                session.clear(); session.update({'user_id': user_row['id'], 'user_type': 'user', 'username': decrypt_data(user_row['cpf'])}); log_missed_punches(user_row['id']); return redirect(url_for('dashboard'))
        flash('Credenciais inválidas.', 'error'); return redirect(url_for('login'))
    return render_template('login.html')

@app.route('/dashboard')
def dashboard():
    if 'username' not in session: return redirect(url_for('login'))
    user_type = session.get('user_type')
    if user_type == 'admin':
        conn = get_db_connection(); users_raw = conn.execute('SELECT id, nome FROM users').fetchall(); conn.close()
        users_list = [{'id': u['id'], 'nome': decrypt_data(u['nome'])} for u in users_raw]
        return render_template('dashboard.html', user_type='admin', users=users_list, active_page='dashboard')
    elif user_type == 'user':
        user_id = session.get('user_id')
        conn = get_db_connection(); user_row = conn.execute('SELECT nome FROM users WHERE id = ?', (user_id,)).fetchone(); schedule_row = conn.execute('SELECT * FROM registros_ponto WHERE user_id = ?', (user_id,)).fetchone(); conn.close()
        user_name = decrypt_data(user_row['nome']) if user_row else "Usuário"; has_schedule = bool(schedule_row)
        available_actions = get_available_actions(user_id) if has_schedule else {}
        return render_template('dashboard.html', user_type='user', user_name=user_name, has_schedule=has_schedule, actions=available_actions)
    return redirect(url_for('login'))

@app.route('/salvar_horario', methods=['POST'])
def salvar_horario():
    if session.get('user_type') != 'admin': return redirect(url_for('login'))
    user_id, e1, s1, e2, s2 = request.form['user_id'], request.form.get('entrada1'), request.form.get('saida1'), request.form.get('entrada2'), request.form.get('saida2')
    enc_e1, enc_s1, enc_e2, enc_s2 = (encrypt_data(v) if v else None for v in [e1, s1, e2, s2])
    conn = get_db_connection()
    if conn.execute('SELECT id FROM registros_ponto WHERE user_id = ?', (user_id,)).fetchone():
        conn.execute('UPDATE registros_ponto SET entrada1=?, saida1=?, entrada2=?, saida2=? WHERE user_id = ?', (enc_e1, enc_s1, enc_e2, enc_s2, user_id))
        flash('Horário atualizado!', 'success')
    else:
        conn.execute('INSERT INTO registros_ponto (user_id, entrada1, saida1, entrada2, saida2) VALUES (?, ?, ?, ?, ?)', (user_id, enc_e1, enc_s1, enc_e2, enc_s2))
        flash('Horário registrado!', 'success')
    conn.commit(); conn.close()
    return redirect(url_for('dashboard'))

@app.route('/register', methods=['POST'])
def register_user():
    if session.get('user_type') != 'admin': return redirect(url_for('login'))
    nome, cpf = request.form['nome'], request.form['cpf']
    try:
        otp_secret = pyotp.random_base32()
        conn = get_db_connection()
        conn.execute('INSERT INTO users (nome, cpf, otp_secret) VALUES (?, ?, ?)', (encrypt_data(nome), encrypt_data(cpf), encrypt_data(otp_secret)))
        conn.commit(); conn.close()
        flash(f'Usuário {nome} cadastrado!', 'success')
        return redirect(url_for('show_qr', cpf=cpf))
    except sqlite3.IntegrityError:
        flash('Este CPF já está cadastrado.', 'error')
    return redirect(url_for('dashboard'))

@app.route('/registrar_meu_ponto/<tipo_registro>', methods=['POST'])
def registrar_meu_ponto(tipo_registro):
    if session.get('user_type') != 'user': return redirect(url_for('login'))
    
    justificativa = request.form.get('justificativa')
    enc_justificativa = encrypt_data(justificativa) if justificativa else None
    
    now = datetime.now(BR_TIMEZONE)
    conn = get_db_connection()
    
    conn.execute('''
        INSERT INTO historico_ponto (user_id, data_registro, hora_registro, tipo_registro, justificativa) 
        VALUES (?, ?, ?, ?, ?)
    ''', (
        session.get('user_id'), 
        encrypt_data(now.strftime('%Y-%m-%d')), 
        encrypt_data(now.strftime('%H:%M:%S')), 
        encrypt_data(tipo_registro),
        enc_justificativa
    ))
    
    conn.commit(); conn.close()
    
    msg_tipo = "com justificativa" if justificativa else ""
    flash(f'Ponto "{tipo_registro.upper().replace("1"," 1")}" registrado às {now:%H:%M:%S} {msg_tipo}!', 'success')
    return redirect(url_for('dashboard'))

@app.route('/admin/update_justification', methods=['POST'])
def update_justification():
    if session.get('user_type') != 'admin': return redirect(url_for('login'))
    record_id, text = request.form.get('record_id'), request.form.get('justification_text')
    if not record_id: flash('Erro ao processar a nota.', 'error'); return redirect(url_for('admin_reports'))
    conn = get_db_connection()
    if text: conn.execute('UPDATE historico_ponto SET justificativa = ? WHERE id = ?', (encrypt_data(text), record_id))
    else: conn.execute('UPDATE historico_ponto SET justificativa = NULL WHERE id = ?', (record_id,))
    conn.commit(); conn.close()
    flash('Nota/Justificativa atualizada!', 'success')
    return redirect(url_for('admin_reports'))

@app.route('/admin/reports', methods=['GET', 'POST'])
def admin_reports():
    if session.get('user_type') != 'admin': return redirect(url_for('login'))
    conn = get_db_connection(); users_raw = conn.execute('SELECT id, nome FROM users').fetchall(); conn.close()
    users_list = [{'id': u['id'], 'nome': decrypt_data(u['nome'])} for u in users_raw]
    years = [datetime.now().year + i for i in range(-2, 1)]; records = None
    filters = session.get('report_filters', {})
    if request.method == 'POST':
        filters = request.form
        session['report_filters'] = filters
    if filters:
        sel_year, sel_month, sel_user_id = int(filters.get('year')), int(filters.get('month')), filters.get('user_id')
        users_map = {u['id']: u['nome'] for u in users_list}; conn = get_db_connection()
        history_raw = conn.execute('SELECT * FROM historico_ponto').fetchall(); conn.close()
        filtered_records = []
        for r in history_raw:
            try:
                if sel_user_id != 'all' and str(r['user_id']) != sel_user_id: continue
                record_date = datetime.strptime(decrypt_data(r['data_registro']), '%Y-%m-%d')
                if record_date.year == sel_year and record_date.month == sel_month:
                    hora = decrypt_data(r['hora_registro']) if r['hora_registro'] else 'AUSENTE'; justification = decrypt_data(r['justificativa']) if r['justificativa'] else None
                    filtered_records.append({'id': r['id'], 'user_name': users_map.get(r['user_id'], 'Desconhecido'),'date': record_date.strftime('%d/%m/%Y'), 'time': hora, 'type': decrypt_data(r['tipo_registro']).replace('1',' 1').replace('2', ' 2').title(), 'justification': justification})
            except Exception: continue
        records = sorted(filtered_records, key=lambda x: (datetime.strptime(x['date'], '%d/%m/%Y'), x['time'] if x['time'] != 'AUSENTE' else '00:00:00'))
    return render_template('reports.html', years=years, records=records, users=users_list, active_page='reports', filters=filters)

def gerar_excel_estilizado(user_name, user_cpf, mes, ano, records):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Ponto_{user_name.split()[0]}"
    
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE 
    

    ws.page_setup.fitToPage = True
    ws.page_setup.fitToHeight = False 
    ws.page_setup.fitToWidth = 1

    ws.print_options.horizontalCentered = True
    
    ws.print_title_rows = '1:8'

    ws.page_margins.left = 0.25
    ws.page_margins.right = 0.25
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5
    ws.page_margins.header = 0.3
    ws.page_margins.footer = 0.3
    
    cor_azul_escuro = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    cor_azul_claro = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    cor_cinza_claro = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    
    fonte_titulo = Font(name='Calibri', size=18, bold=True, color="FFFFFF")
    fonte_negrito = Font(name='Calibri', size=11, bold=True)
    fonte_normal = Font(name='Calibri', size=11)
    
    borda_fina = Side(border_style="thin", color="000000")
    border_box = Border(left=borda_fina, right=borda_fina, top=borda_fina, bottom=borda_fina)

    ws.merge_cells('A1:F2')
    cell_titulo = ws['A1']
    cell_titulo.value = "FOLHA DE PONTO INDIVIDUAL"
    cell_titulo.font = fonte_titulo
    cell_titulo.alignment = Alignment(horizontal='center', vertical='center')
    cell_titulo.fill = cor_azul_escuro
    
    ws.merge_cells('A4:F4')
    ws['A4'].value = "DADOS DO COLABORADOR"
    ws['A4'].font = fonte_negrito
    ws['A4'].fill = cor_azul_claro
    ws['A4'].border = border_box

    ws.merge_cells('A5:B5'); ws.merge_cells('C5:F5')
    ws['A5'] = "Nome:"; ws['A5'].font = fonte_negrito
    ws['C5'] = user_name; ws['C5'].font = fonte_normal
    
    ws.merge_cells('A6:B6'); ws.merge_cells('C6:D6')
    ws['A6'] = "CPF:"; ws['A6'].font = fonte_negrito
    ws['C6'] = user_cpf; ws['C6'].font = fonte_normal
    
    ws['E6'] = "Período:"; ws['E6'].font = fonte_negrito
    ws['F6'] = f"{mes:02d}/{ano}"; ws['F6'].font = fonte_normal

    for row in ws['A5:F6']:
        for cell in row:
            cell.border = border_box
            cell.alignment = Alignment(vertical='center')

    col_titles = ['Data', 'Entrada 1', 'Saída 1', 'Entrada 2', 'Saída 2', 'Observação']
    ws.append([]) 
    ws.append(col_titles) 
    
    for col_num, cell in enumerate(ws[8], 1):
        cell.font = fonte_negrito
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = cor_azul_claro
        cell.border = border_box

    day_records = {}
    for rec in records:
        day = rec['date']
        if day not in day_records: day_records[day] = {'Data': day}
        punch_map = {'Entrada 1': 'Entrada 1', 'Saida 1': 'Saída 1', 'Entrada 2': 'Entrada 2', 'Saida 2': 'Saída 2'}
        
        tipo_formatado = punch_map.get(rec['type'])
        if tipo_formatado:
            day_records[day][tipo_formatado] = rec['time']
        
        if rec['justification']:
            prev_obs = day_records[day].get('Observação', '')
            day_records[day]['Observação'] = f"{prev_obs}{rec['type']}: {rec['justification']}\n".strip()

    row_idx = 9
    for day_data in sorted(day_records.values(), key=lambda x: datetime.strptime(x['Data'], '%d/%m/%Y')):
        row_data = [day_data.get(title, '') for title in col_titles]
        ws.append(row_data)
        
        ws.row_dimensions[row_idx].height = 40 
        
        for col_num in range(1, 7):
            cell = ws.cell(row=row_idx, column=col_num)
            cell.border = border_box
            cell.font = fonte_normal
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
            if col_num == 6: 
                cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            
            if row_idx % 2 == 0:
                cell.fill = cor_cinza_claro
                
        row_idx += 1

    ws.column_dimensions['A'].width = 15 
    ws.column_dimensions['B'].width = 13
    ws.column_dimensions['C'].width = 13
    ws.column_dimensions['D'].width = 13
    ws.column_dimensions['E'].width = 13
    ws.column_dimensions['F'].width = 75 

    sig_start_row = row_idx + 3
    
    ws.merge_cells(f'A{sig_start_row}:C{sig_start_row}')
    sig_cell_emp = ws[f'A{sig_start_row}']
    sig_cell_emp.value = user_name
    sig_cell_emp.font = fonte_normal
    sig_cell_emp.alignment = Alignment(horizontal='center', vertical='bottom')
    sig_cell_emp.border = Border(top=borda_fina)
    
    ws.merge_cells(f'A{sig_start_row+1}:C{sig_start_row+1}')
    label_emp = ws[f'A{sig_start_row+1}']
    label_emp.value = "Assinatura do Colaborador"
    label_emp.alignment = Alignment(horizontal='center', vertical='top')
    label_emp.font = Font(name='Calibri', size=9, italic=True)

    ws.merge_cells(f'D{sig_start_row}:F{sig_start_row}')
    sig_cell_boss = ws[f'D{sig_start_row}']
    sig_cell_boss.value = "Visto do Empregador"
    sig_cell_boss.font = fonte_normal
    sig_cell_boss.alignment = Alignment(horizontal='center', vertical='bottom')
    sig_cell_boss.border = Border(top=borda_fina)
    
    ws.merge_cells(f'D{sig_start_row+1}:F{sig_start_row+1}')
    label_boss = ws[f'D{sig_start_row+1}']
    label_boss.value = f"Data: {datetime.now().strftime('%d/%m/%Y')}"
    label_boss.alignment = Alignment(horizontal='center', vertical='top')
    label_boss.font = Font(name='Calibri', size=9, italic=True)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

@app.route('/admin/export_report', methods=['POST'])
def export_report():
    if session.get('user_type') != 'admin': return redirect(url_for('login'))
    
    sel_year = int(request.form.get('year'))
    sel_month = int(request.form.get('month'))
    sel_user_id = request.form.get('user_id')
    
    if sel_user_id == 'all':
        flash('Por favor, selecione um usuário específico para exportar o recibo.', 'error')
        return redirect(url_for('admin_reports'))
        
    conn = get_db_connection()
    user_data = conn.execute('SELECT nome, cpf FROM users WHERE id = ?', (sel_user_id,)).fetchone()
    
    if not user_data: 
        conn.close()
        flash('Usuário selecionado não encontrado.', 'error')
        return redirect(url_for('admin_reports'))
        
    user_name = decrypt_data(user_data['nome'])
    user_cpf = decrypt_data(user_data['cpf'])
    
    history_raw = conn.execute('SELECT * FROM historico_ponto WHERE user_id = ?', (sel_user_id,)).fetchall()
    conn.close()
    
    records = []
    for r in history_raw:
        try:
            record_date = datetime.strptime(decrypt_data(r['data_registro']), '%Y-%m-%d')
            if record_date.year == sel_year and record_date.month == sel_month:
                records.append({
                    'date': record_date.strftime('%d/%m/%Y'), 
                    'time': decrypt_data(r['hora_registro']) if r['hora_registro'] else 'AUSENTE', 
                    'type': decrypt_data(r['tipo_registro']).replace('1',' 1').replace('2', ' 2').title(), 
                    'justification': decrypt_data(r['justificativa']) if r['justificativa'] else ''
                })
        except Exception: continue
    
    buffer = gerar_excel_estilizado(user_name, user_cpf, sel_month, sel_year, records)
    
    filename = f"Recibo_Ponto_{user_name.replace(' ','_')}_{sel_month:02d}_{sel_year}.xlsx"
    return send_file(
        buffer, 
        as_attachment=True, 
        download_name=filename, 
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@app.route('/export_meu_ponto')
def export_meu_ponto():
    if session.get('user_type') != 'user': return redirect(url_for('login'))
    
    user_id = session.get('user_id')
    now = datetime.now(BR_TIMEZONE)
    sel_month, sel_year = now.month, now.year
    
    conn = get_db_connection()
    user_data = conn.execute('SELECT nome, cpf FROM users WHERE id = ?', (user_id,)).fetchone()
    
    if not user_data: 
        conn.close()
        flash('Erro ao recuperar dados do usuário.', 'error')
        return redirect(url_for('dashboard'))
        
    user_name = decrypt_data(user_data['nome'])
    user_cpf = decrypt_data(user_data['cpf'])
    
    history_raw = conn.execute('SELECT * FROM historico_ponto WHERE user_id = ?', (user_id,)).fetchall()
    conn.close()
    
    records = []
    for r in history_raw:
        try:
            record_date = datetime.strptime(decrypt_data(r['data_registro']), '%Y-%m-%d')
            if record_date.year == sel_year and record_date.month == sel_month:
                records.append({
                    'date': record_date.strftime('%d/%m/%Y'), 
                    'time': decrypt_data(r['hora_registro']) if r['hora_registro'] else 'AUSENTE', 
                    'type': decrypt_data(r['tipo_registro']).replace('1',' 1').replace('2', ' 2').title(), 
                    'justification': decrypt_data(r['justificativa']) if r['justificativa'] else ''
                })
        except Exception: continue
            
    buffer = gerar_excel_estilizado(user_name, user_cpf, sel_month, sel_year, records)
    
    filename = f"Meu_Recibo_{sel_month:02d}_{sel_year}.xlsx"
    return send_file(
        buffer, 
        as_attachment=True, 
        download_name=filename, 
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@app.route('/qr_code/<cpf>')
def show_qr(cpf):
    if session.get('user_type') != 'admin': return redirect(url_for('login'))
    conn = get_db_connection(); users_db = conn.execute('SELECT * FROM users').fetchall(); conn.close()
    for row in users_db:
        if decrypt_data(row['cpf']) == cpf:
            nome_usuario, otp_secret = decrypt_data(row['nome']), decrypt_data(row['otp_secret'])
            otp_uri = pyotp.totp.TOTP(otp_secret).provisioning_uri(name=f"{nome_usuario} ({cpf})", issuer_name="fPontoBhsoft")
            img = qrcode.make(otp_uri); buf = BytesIO(); img.save(buf); qr_code_img = base64.b64encode(buf.getvalue()).decode('ascii')
            return render_template('qr_display.html', user_name=nome_usuario, qr_code_img=qr_code_img)
    flash('Usuário não encontrado.', 'error'); return redirect(url_for('dashboard'))

@app.route('/logout')
def logout():
    session.clear(); flash('Você foi desconectado com sucesso.', 'success'); return redirect(url_for('login'))

if __name__ == '__main__':
    app.run(host="0.0.0.0", port=80, debug=True)
